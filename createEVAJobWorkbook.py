'''
Program to take in a .xslx for the estimated cost detail, actual cost detail, and Revenue report workbook file and fill out all Job Cost Report Sheets.

Usage:  

    python createEVAJobWorkbook.py path_to_eva_workbook_file 

The processed total job workbook will be saved as a copy in /processed

Requires openpyxl

    pip install openpyxl

Author: Brian Wright
9-6-2022

'''
from xmlrpc.client import MAXINT
from util import draw_line, set_border, copySheet

import sys
import openpyxl
from openpyxl.styles import Font
import os 
import shutil
from copy import copy
from datetime import datetime
from collections import OrderedDict
       

def createEVAJobWorkbook(eva_total_wb_path):
    # copy wb and work on copy
    processed_file_path = os.path.split(eva_total_wb_path)[0]  +'/processed/' + os.path.basename(eva_total_wb_path).split('.')[0] + '_processed.xlsx'
    shutil.copyfile(eva_total_wb_path, processed_file_path)

    eva_total_wb = openpyxl.load_workbook(processed_file_path) 
    if not eva_total_wb:
        print("Error: failed to open workbook: ", processed_file_path)
        return

    actual_cost_detail_sheet    = eva_total_wb.worksheets[0]
    revenue_sheet               = eva_total_wb.worksheets[1]
    estimate_cost_detail_sheet  = eva_total_wb.worksheets[2]

    #job_str_set = set()
    job_str_set = OrderedDict()

    ACTUAL_NAME_COLUMN   = 11      
    REVENUE_NAME_COLUMN  = 11
    ESTIMATE_NAME_COLUMN = 10

    # add new sheet for each unique job, aggregating from all sheets as a precaution
    # column
    for i in range(1, actual_cost_detail_sheet.max_row + 1): 
        job_data = actual_cost_detail_sheet.cell(row = i, column = ACTUAL_NAME_COLUMN).value

        # format is currrently:   job_name:job_number type
        # is a job string? 
        if job_data and len(job_data.split(":")) > 1:
            job_number = job_data.split(":")[1].split(' ')[0]    # could be better?
            job_str_set[job_number] = None

    for i in range(1, revenue_sheet.max_row + 1): 
        job_data = revenue_sheet.cell(row = i, column = REVENUE_NAME_COLUMN).value

        if job_data and len(job_data.split(":")) > 1:
            job_number = job_data.split(":")[1].split(' ')[0]    # could be better?
            job_str_set[job_number] = None

    for i in range(1, estimate_cost_detail_sheet.max_row + 1): 
        job_data = revenue_sheet.cell(row = i, column = ESTIMATE_NAME_COLUMN).value

        if job_data and len(job_data.split(":")) > 1:
            job_number = job_data.split(":")[1].split(' ')[0]    # could be better?
            job_str_set[job_number] = None

    job_numbers = list(job_str_set.keys())
    
    if len(job_numbers) == 0:
        print("Error: failed to find jobs in workbook: ", processed_file_path)
        return
    
    # add new sheet for each job number
    for job_number in job_numbers:
        # clear old sheet
        if (job_number in eva_total_wb.sheetnames):
            eva_total_wb.remove(eva_total_wb[job_number])
        eva_total_wb.create_sheet(title=job_number)

    # copy empty job cost sheet
    eva_jc_wb = openpyxl.load_workbook(os.getcwd() + "/data/eva_jc_blank.xlsx") 
    if not eva_jc_wb:
        print("Error: failed to open data workbook: /data/eva_jc_blank.xlsx")
        return

    ACTUAL_DATE_COLUMN = 9
    ACTUAL_ITEM_COLUMN = 15      
    ACTUAL_AMOUNT_COLUMN = 19

    class JobItem:
        def __init__(self, item_name="", sub=False):
            self.item_name = item_name
            if "Income" in self.item_name:
                self._job_number = MAXINT # want income to be last in the order.
            else:
                self._job_number = int(item_name.split(" ")[0]) # catches them all currently... can be "Income"
            self.actual_amount   = 0  # used for tracking value for non sub-type
            self.estimate_amount = 0  

            self.hasSub = sub 
            self.actual_sub_items   = OrderedDict() 
            self.estimate_sub_items = OrderedDict() # contains list of (name, amount) pairs

        
        def processSubItem(self, sub_item_name, amount, actual=False):
            if not self.hasSub:
                print("error: tried to process subitem on top level item")
                return

            if sub_item_name not in self.actual_sub_items.keys():
                # initialize sub-item in both dicts
                if actual:
                    self.actual_sub_items[sub_item_name] = amount 
                    self.estimate_sub_items[sub_item_name] = 0
                else:
                    self.actual_sub_items[sub_item_name] = 0 
                    self.estimate_sub_items[sub_item_name] = amount 
            else: # already in?
                if actual:
                    self.actual_sub_items[sub_item_name] += amount
                else:
                    self.estimate_sub_items[sub_item_name] += amount
        
        def getActualTotal(self):
            if self.hasSub:
                actual_total = 0
                for k,v in self.actual_sub_items:
                    actual_total += v
                return actual_total

            return self.actual_amount

        def getEstimateTotal(self):
            if self.hasSub:
                estimate_total = 0
                for k,v in self.estimate_sub_items:
                    estimate_total += v
                return estimate_total

            return self.estimate_total

        def __lt__(self, other_job: object) -> bool:
           return self._job_number < other_job._job_number
            
    # SCOPED
    def createEVAJobCostSheet(sheet):
        # create job sheet
        job_number = sheet.title

        if not job_number:
            print("Warn: null sheet passed")
            return

        min_date = datetime.max
        max_date = datetime.min

        job_name = None
        job_items = []
        # get all job actual costs 
        for i in range(1, actual_cost_detail_sheet.max_row + 1):    # could optimize by not doing all rows
            j_name = actual_cost_detail_sheet.cell(row = i, column = ACTUAL_NAME_COLUMN).value

            if j_name and job_number in j_name:
                
                if not job_name:
                    job_name = j_name 

                j_item = actual_cost_detail_sheet.cell(row = i, column = ACTUAL_ITEM_COLUMN).value
                j_actual_amount = actual_cost_detail_sheet.cell(row = i, column = ACTUAL_AMOUNT_COLUMN).value
                if not j_item:
                    print("Warn: have job entry with no item data", j_name)
                    continue

                if not j_actual_amount:
                    #print("Warn: have job entry with no actual amount, ", j_name)
                    continue

                date = actual_cost_detail_sheet.cell(row = i, column = ACTUAL_DATE_COLUMN).value
                if date:
                    min_date = min(min_date, date)
                    max_date = max(max_date, date)
                else:
                    print("Warn: Actual job without a date: ", j_name)


                item_name = ""
                sub_item_name = None
                # won't have sub item types without :
                if ":" not in j_item:
                    item_name = j_item
                elif ":" in j_item:
                    item_name, sub_item_name = j_item.split(":")
                else:
                    print("Warn: unhandled job item: ", j_item)

                job_item = None
                # find job_item if it already exists
                for j_item in job_items:
                    if j_item.item_name == item_name:
                        job_item = j_item
                if not job_item:
                    if sub_item_name:
                        job_item = JobItem(item_name, True)
                        job_item.processSubItem(sub_item_name, j_actual_amount, actual=True)
                    else:
                        job_item = JobItem(item_name, False)
                        job_item.actual_amount += j_actual_amount
                    job_items.append(job_item)
                else: # 
                    if job_item.hasSub:
                        job_item.processSubItem(sub_item_name, j_actual_amount, actual=True)
                    else:
                        job_item.actual_amount += j_actual_amount

        ESTIMATE_ITEM_COLUMN   = 12
        ESTIMATE_AMOUNT_COLUMN = 16

        # now get estimate amounts for all the job items, should not be any new jobs
        for i in range(1, estimate_cost_detail_sheet.max_row + 1):    # could optimize by not doing all rows
            j_name = estimate_cost_detail_sheet.cell(row = i, column = ESTIMATE_NAME_COLUMN).value

            if j_name and job_number in j_name:

                if not job_name:
                    job_name = j_name 

                j_item = estimate_cost_detail_sheet.cell(row = i, column = ESTIMATE_ITEM_COLUMN).value
                j_estimate_amount = estimate_cost_detail_sheet.cell(row = i, column = ESTIMATE_AMOUNT_COLUMN).value

                if not j_item:
                    print("Warn: have job entry with no item data", j_name)
                    continue

                if not j_estimate_amount:
                    print("Warn: have job entry with no estimate amount, ", j_name)
                    continue

                item_name = ""
                sub_item_name = None
                # won't have sub item types without :
                if ":" not in j_item:
                    item_name = j_item
                elif ":" in j_item:
                    item_name, sub_item_name = j_item.split(":")
                else:
                    print("Warn: unhandled job item: ", j_item)

                job_item = None
                # find job_item if it already exists
                for j_item in job_items:
                    if j_item.item_name == item_name:
                        job_item = j_item
                if not job_item:
                    #print("Warn: found job with an estimate but no actual cost, ", item_name, " ", sub_item_name)
                    job_item = JobItem(item_name)
                    if sub_item_name:
                        job_item.hasSub = True
                        job_item.processSubItem(sub_item_name, j_estimate_amount, actual=False)
                    else:
                        job_item.estimate_amount += j_estimate_amount
                    job_items.append(job_item)
                else:
                    if job_item.hasSub:
                        job_item.processSubItem(sub_item_name, j_estimate_amount, actual=False)
                    else:
                        job_item.estimate_amount += j_estimate_amount
                
        # last effort to get job name
        if not job_name:
            for i in range(1, revenue_sheet.max_row + 1):    # could optimize by not doing all rows
                j_name = revenue_sheet.cell(row = i, column = REVENUE_NAME_COLUMN).value
                if j_name and job_number in j_name:
                    job_name = j_name 
                    break

        # append job name at top text
        if not job_name:
            print("Warn couldn't get job name job number: ", job_number)
            return

        sheet.cell(row = 2, column = 1).value = sheet.cell(row = 2, column = 1).value + " " + job_name

        ITEM_NAME_COLUMN        = 3
        SUBITEM_NAME_COLUMN     = 4
        ESTIMATE_COST_COLUMN    = 5
        ACT_COST_COLUMN         = 7
        DIFF_COLUMN             = 9

        # date and time
        sheet.cell(row = 1, column = DIFF_COLUMN).value = datetime.today().strftime("%H:%M %p")
        sheet.cell(row = 2, column = DIFF_COLUMN).value = datetime.today().strftime("%B %d, %Y")

        i = 7  # starting point after 'Service' row

        total_actual_labor_cost    = 0
        total_estimate_labor_cost  = 0
        total_labor_cost_no_temp   = 0
        total_temp_labor_cost      = 0

        total_estimate_cost = 0
        total_actual_cost   = 0

        job_items.sort()

        FONT = "Arial"
        FONT_SIZE = 9
        default_font = Font(bold=False,name=FONT,sz=FONT_SIZE)
        bold_font = Font(bold=True,name=FONT,sz=FONT_SIZE)

        # write job cost data
        for job_item in job_items:
            if not job_item.hasSub:

                if "labor" in job_item.item_name.lower() and "sub" not in job_item.item_name.lower():
                    total_actual_labor_cost += job_item.actual_amount
                    total_estimate_labor_cost += job_item.estimate_amount
                    if "temp" not in job_item.item_name.lower():
                        total_labor_cost_no_temp += job_item.actual_amount
                    elif "temp" in job_item.item_name.lower():
                        total_temp_labor_cost += job_item.actual_amount

                total_actual_cost += job_item.actual_amount
                total_estimate_cost += job_item.estimate_amount
                diff = job_item.estimate_amount - job_item.actual_amount

                sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = job_item.item_name 
                sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = job_item.estimate_amount
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
                sheet.cell(row = i, column = ACT_COST_COLUMN).value = job_item.actual_amount
                sheet.cell(row = i, column = ACT_COST_COLUMN).font = bold_font
                sheet.cell(row = i, column = DIFF_COLUMN).value = diff 
                sheet.cell(row = i, column = DIFF_COLUMN).font = bold_font
                i += 1
                # Draw line 
                cell_range = "E" + str(i-1) + ":I" + str(i-1)
                draw_line(sheet, cell_range) 

            else:
                sheet.cell(row = i, column = 3).value = job_item.item_name 
                i += 1
                sub_actual_total = 0
                sub_estimate_total = 0
                for s_item_name, s_actual_amount in job_item.actual_sub_items.items():
                    
                    s_estimate_amount = job_item.estimate_sub_items[s_item_name] 

                    if "labor" in s_item_name.lower() and "sub" not in s_item_name.lower():
                        total_actual_labor_cost += s_actual_amount
                        total_estimate_labor_cost += s_estimate_amount
                        if "temp" not in s_item_name.lower():
                            total_labor_cost_no_temp += s_actual_amount
                        elif "temp" in s_item_name.lower():
                            total_temp_labor_cost += s_actual_amount


                    total_actual_cost += s_actual_amount
                    total_estimate_cost += s_estimate_amount
                    #TODO: could handle these in the job items themselves
                    sub_actual_total += s_actual_amount
                    sub_estimate_total += s_estimate_amount
                    s_diff = s_estimate_amount - s_actual_amount

                    sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = s_item_name
                    sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = s_estimate_amount
                    sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = default_font
                    sheet.cell(row = i, column = ACT_COST_COLUMN).value = s_actual_amount
                    sheet.cell(row = i, column = ACT_COST_COLUMN).font = default_font
                    sheet.cell(row = i, column = DIFF_COLUMN).value = s_diff
                    sheet.cell(row = i, column = DIFF_COLUMN).font = default_font
                    i += 1
                # write out total for the subs
                sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total " + job_item.item_name
                sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = sub_estimate_total
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
                sheet.cell(row = i, column = ACT_COST_COLUMN).value = sub_actual_total
                sheet.cell(row = i, column = ACT_COST_COLUMN).font = bold_font
                sheet.cell(row = i, column = DIFF_COLUMN).value = sub_estimate_total - sub_actual_total
                sheet.cell(row = i, column = DIFF_COLUMN).font = bold_font
                i += 1
                # Draw line 
                cell_range = "E" + str(i-2) + ":I" + str(i-2)
                draw_line(sheet, cell_range) 

        # whitespace
        i += 1

        # total
        sheet.cell(row = i, column = 2).value = "Total"
        sheet.cell(row = i, column = 2).font = Font(bold=True)
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_estimate_cost
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_actual_cost
        sheet.cell(row = i, column = DIFF_COLUMN).value = total_estimate_cost - total_actual_cost
        # total font bold
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = bold_font
        sheet.cell(row = i, column = DIFF_COLUMN).font = bold_font
        i += 1
        # whitespace
        i += 1

        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Actual Revenue To Date"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        i += 1

        DATE_REVENUE_COLUMN = 9
        NAME_REVENUE_COLUMN = 11
        MEMO_REVENUE_COLUMN = 13
        ITEM_REVENUE_COLUMN = 15
        AMOUNT_REVENUE_COLUMN = 19
 
        # get revenue info 

        total_orig_contract    = 0
        total_change_order     = 0
        total_other_job_income = 0
        total_retainage        = 0

        for j in range(6, revenue_sheet.max_row + 1):    
            j_name = revenue_sheet.cell(row = j, column = NAME_REVENUE_COLUMN).value
            if j_name and job_number in j_name:
                
                amount = revenue_sheet.cell(row = j, column = AMOUNT_REVENUE_COLUMN).value
                item_str = revenue_sheet.cell(row = j, column = ITEM_REVENUE_COLUMN).value
                if not item_str:
                    continue
                item_str = item_str.lower()

                # issue with the dates
                # https://stackoverflow.com/questions/57530408/python-reads-inaccurately-from-excel-date-time-field
                '''
                date = str(revenue_sheet.cell(row = i, column = DATE_REVENUE_COLUMN).value)
                if date:
                    min_date = min(min_date, date)
                    max_date = max(max_date, date)
                else:
                    print("Warn: Revenue Job without a date: ", j_name)
                '''

                if "orig contract" in item_str:
                    total_orig_contract += amount
                elif "change order" in item_str:
                    total_change_order += amount
                elif "other job income" in item_str:
                    total_other_job_income += amount
                elif "retainage" in item_str:
                    total_retainage += amount
                
            total_billed_before_retainage = total_orig_contract + total_change_order + total_other_job_income



        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Orig Contract"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_orig_contract
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Change Order"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_change_order 
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Other Job Income"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_other_job_income 
        i += 1
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total Billed to Date"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_billed_before_retainage
        i += 1
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Retainage Held by Customer"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_retainage
        i += 1
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total Actual Revenue Collected to Date"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_billed_before_retainage + total_retainage
        i += 1
        i += 1 # whitespace

        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total Estimated Labor including Temp Labor"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_estimate_labor_cost
        i += 1
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total Actual Labor including Temp Labor"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_labor_cost_no_temp + .3 * total_labor_cost_no_temp + total_temp_labor_cost 
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Estimated vs Actual Difference"
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_estimate_labor_cost - (total_labor_cost_no_temp + .3 * total_labor_cost_no_temp + total_temp_labor_cost)
        i += 1
        sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Percent Complete based on Labor Costs"
        if total_estimate_labor_cost > 0:
            sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = str(round((total_actual_labor_cost/total_estimate_labor_cost) * 100, 2)) + "%"
        else:
            sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = "0.0%"
        i += 1
        i += 1 # whitespace

    
        # summary box
        '''
        Calculation details for the summary box:
        Total Labor: Add all labor costs except temp labor
        Labor OH: Multiply 30% to the total labor calculated above
        Other OH: Multiply 0.005 to all costs
        Total Cost w/OH: Total Costs + Labor OH + Other OH
        '''
        labor_oh = total_labor_cost_no_temp * 0.3
        other_oh = total_actual_cost * 0.005
        total_cost_w_oh = total_actual_cost + labor_oh + other_oh

        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total Labor"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_labor_cost_no_temp 
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Labor OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = labor_oh
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Other OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = other_oh
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total Cost w/ OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_cost_w_oh
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Billed To Date"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_billed_before_retainage 
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1

        # column letter row number : column letter row number  for top left, bottom right
        cell_range = "D" + str(i-5) + ":E" + str(i-1)
        set_border(sheet, cell_range)

        # whitespace
        i += 1

        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Billed To Date"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        i += 1

        # Grab all billed
        # Could avoid doing this iteration twice
        for j in range(6, revenue_sheet.max_row + 1):    
            j_name = revenue_sheet.cell(row = j, column = NAME_REVENUE_COLUMN).value
            if j_name and job_number in j_name:
                memo_cell = revenue_sheet.cell(row = j, column = MEMO_REVENUE_COLUMN) 
                amount_cell = revenue_sheet.cell(row = j, column = AMOUNT_REVENUE_COLUMN) 

                sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = memo_cell.value
                sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = default_font
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = amount_cell.value
                sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = default_font
                i += 1

        # write total billed to date items
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total" 
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = bold_font
        #NOTE: should be equivalent, could calc again to be safe
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).value = total_billed_before_retainage + total_retainage
        sheet.cell(row = i, column = ESTIMATE_COST_COLUMN).font = bold_font
        i += 1

        # write date range
        sheet.cell(row = 3, column = 1).value = "Transactions from: " + min_date.strftime("%m/%d/%y") + " to " + max_date.strftime("%m/%d/%y")

    # -------------------------------------------------------------------------------- #

    # create and fill all job sheet data
    # skip first 3
    for i in range(3, len(eva_total_wb.sheetnames)):
        # skip 3 sheets
        print("Processing job: ", i-2, " out of ", len(eva_total_wb.sheetnames)-3)
        sheet = eva_total_wb.worksheets[i]
       # copy initial format into empty sheet
        copySheet(eva_jc_wb.active, sheet)
        createEVAJobCostSheet(sheet)
        sheet.orientation = 'portrait'

    eva_total_wb.save(processed_file_path)


def main(argv):
    if len(argv) == 0 or len(argv) > 1:
        print("Error - usage: supply cost detail workbook , revenue workbook, and actual cost workbook")
        return

    eva_total_wb_path = os.path.abspath(argv[0])

    if os.path.isfile(eva_total_wb_path):
        createEVAJobWorkbook(eva_total_wb_path) 
    elif not os.path.isfile(eva_total_wb_path):
        print("Error: eva path does not exist?: ", eva_total_wb_path)
    else:
        print("Error: wrong input")

if __name__ == "__main__":
   main(sys.argv[1:])