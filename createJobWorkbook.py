'''

Program to take in a .xslx for the cost detail workbook and revenue report workbook file and fill out all Job Cost Report Sheets.

Usage:  

    python createJobWorkbook.py path_to_job_workbook_file   path_to_revenue_report_workbook-file

The processed total job workbook will be saved as a copy in /processed

Requires openpyxl

    pip install openpyxl

Author: Brian Wright
8-18-2022

'''

from util import set_border, copySheet, draw_line

import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import os 
import shutil
from copy import copy
from datetime import datetime
from collections import OrderedDict

def createJobWorkbook(cost_detail_wb_path, revenue_file_path):
    # copy wb and work on copy
    processed_file_path = os.path.split(cost_detail_wb_path)[0]  +'/processed/' + os.path.basename(cost_detail_wb_path).split('.')[0] + '_processed.xlsx'
    shutil.copyfile(cost_detail_wb_path, processed_file_path)

    cost_detail_wb = openpyxl.load_workbook(processed_file_path) 
    if not cost_detail_wb:
        print("Error: failed to open workbook: ", processed_file_path)
        return

    cost_detail_sheet = cost_detail_wb.active
    cost_detail_sheet.title = "Total" # rename

    revenue_wb = openpyxl.load_workbook(revenue_file_path) 
    if not cost_detail_wb:
        print("Error: failed to open workbook: ", revenue_file_path)
        return
    
    revenue_sheet = revenue_wb.active

    #job_str_set = set()
    job_str_set = OrderedDict()
    NAME_COLUMN = 8      # might switch between eva and non eva?

    # add new sheet for each unique job
    # column
    for i in range(1, cost_detail_sheet.max_row + 1): 
        job_data = cost_detail_sheet.cell(row = i, column = NAME_COLUMN).value

        # format is currrently:   job_name:job_number type
        # is a job string? 
        if job_data and len(job_data.split(":")) > 1:
            job_number = job_data.split(":")[1].split(' ')[0]    # could be better?
            job_str_set[job_number] = None

    job_numbers = list(job_str_set.keys())
    
    if len(job_numbers) == 0:
        print("Error: failed to find jobs in workbook: ", processed_file_path)
        return


    # add new sheet for each job number
    for job_number in job_numbers:
        cost_detail_wb.create_sheet(title=job_number)

    # copy empty job cost sheet
    jc_wb = openpyxl.load_workbook(os.getcwd() + "/data/jc_blank.xlsx") 
    if not jc_wb:
        print("Error: failed to open data workbook: /data/jc_blank.xlsx")
        sys.exit()
        return

    DATE_COLUMN = 6
    ITEM_COLUMN = 10      
    AMOUNT_COLUMN = 16

    class JobItem:
        def __init__(self, item_name=""):
            self.item_name = item_name
            self.amount = 0  # used for tracking value for non sub-type

            self.hasSub = False
            self.sub_items = OrderedDict() # contains list of (name, amount) pairs
        
        def processSubItem(self, sub_item_name, amount):
            if not self.hasSub:
                print("error: tried to process subitem on top level item")
                return

            if sub_item_name not in self.sub_items.keys():
                self.sub_items[sub_item_name] = amount 
            else:
                self.sub_items[sub_item_name] += amount
            
    # SCOPED
    def createJobCostSheet(sheet):
        # create job sheet
        job_number = sheet.title

        min_date = datetime.max
        max_date = datetime.min

        job_name = None
        job_items = []
        # get all job progress entries 
        for i in range(1, cost_detail_sheet.max_row + 1):    # could optimize by not doing all rows
            j_name = cost_detail_sheet.cell(row = i, column = NAME_COLUMN).value

            if j_name and job_number in j_name:

                date = cost_detail_sheet.cell(row = i, column = DATE_COLUMN).value
                if date:
                    min_date = min(min_date, date)
                    max_date = max(max_date, date)
                else:
                    print("Warn: Job without a date: ", j_name)

                if not job_name:
                    job_name = j_name 

                j_item = cost_detail_sheet.cell(row = i, column = ITEM_COLUMN).value
                j_amount = cost_detail_sheet.cell(row = i, column = AMOUNT_COLUMN).value
                if not j_item:
                    print("Warn: have job entry with no item data", j_name)
                    continue

                if not j_amount:
                    print("Warn: have job entry with no amount, ", j_name)
                    continue

                item_name = ""
                sub_item_name = None
                # won't have sub item types without :
                if ":" not in j_item:
                    item_name = j_item
                elif ":" in j_item:
                    item_name, sub_item_name = j_item.split(":")
                else:
                    print("Warn: unhandled job item: ", j_item)

                job_item = None
                # find job_item if it already exists
                for j_item in job_items:
                    if j_item.item_name == item_name:
                        job_item = j_item
                if not job_item:
                    job_item = JobItem(item_name)
                    if sub_item_name:
                        job_item.hasSub = True
                        job_item.processSubItem(sub_item_name, j_amount)
                    else:
                        job_item.amount += j_amount
                    job_items.append(job_item)
                else: # 
                    if job_item.hasSub:
                        job_item.processSubItem(sub_item_name, j_amount)
                    else:
                        job_item.amount += j_amount
                        
        # append job name at top text
        sheet.cell(row = 2, column = 1).value = sheet.cell(row = 2, column = 1).value + " " + job_name


        ITEM_NAME_COLUMN    = 3
        SUBITEM_NAME_COLUMN = 4
        ACT_COST_COLUMN     = 5
        ACT_REVENUE_COLUMN  = 7
        DIFF_COLUMN         = 9

        # date and time
        sheet.cell(row = 1, column = DIFF_COLUMN).value = datetime.today().strftime("%H:%M %p")
        sheet.cell(row = 2, column = DIFF_COLUMN).value = datetime.today().strftime("%B %d, %Y")

        i = 7  # starting point after 'Service' row

        total_labor_cost = 0
        total_cost = 0
        # write job cost data
        for job_item in job_items:
            if not job_item.hasSub:

                if "labor" in job_item.item_name.lower() and "temp" not in job_item.item_name.lower():
                    total_labor_cost += job_item.amount

                total_cost += job_item.amount
                sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = job_item.item_name 
                sheet.cell(row = i, column = ACT_COST_COLUMN).value = job_item.amount
                sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = 0.0
                sheet.cell(row = i, column = DIFF_COLUMN).value = -job_item.amount
                i += 1
                # Draw line 
                cell_range = "E" + str(i-1) + ":I" + str(i-1)
                draw_line(sheet, cell_range) 

            else:
                sheet.cell(row = i, column = 3).value = job_item.item_name 
                i += 1
                sub_total = 0
                for s_item_name, s_amount in job_item.sub_items.items():

                    if "labor" in s_item_name.lower() and "temp" not in s_item_name.lower():
                        total_labor_cost += s_amount 

                    total_cost += s_amount
                    sub_total += s_amount
                    sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = s_item_name
                    sheet.cell(row = i, column = ACT_COST_COLUMN).value = s_amount
                    sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = 0.0
                    sheet.cell(row = i, column = DIFF_COLUMN).value = -s_amount
                    i += 1
                # write out total for the subs
                sheet.cell(row = i, column = ITEM_NAME_COLUMN).value = "Total " + job_item.item_name
                sheet.cell(row = i, column = ACT_COST_COLUMN).value = sub_total
                sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = 0.0
                sheet.cell(row = i, column = DIFF_COLUMN).value = -sub_total
                i += 1
                # Draw line 
                cell_range = "E" + str(i-2) + ":I" + str(i-2)
                draw_line(sheet, cell_range) 

        DATE_REVENUE_COLUMN = 9
        NUM_REVENUE_COLUMN = 11
        NAME_REVENUE_COLUMN = 13
        MEMO_REVENUE_COLUMN = 15
        ITEM_REVENUE_COLUMN = 17
        AMOUNT_REVENUE_COLUMN = 19
 
        # get total income
        total_revenue_income = 0
        for j in range(6, revenue_sheet.max_row + 1):    
            j_name = revenue_sheet.cell(row = j, column = NAME_REVENUE_COLUMN).value
            if j_name and job_number in j_name:

                '''
                date = revenue_sheet.cell(row = i, column = DATE_REVENUE_COLUMN).value
                if date:
                    min_date = min(min_date, date)
                    max_date = max(max_date, date)
                else:
                    print("Warn: Job without a date: ", j_name)
                '''

                amount_cell = revenue_sheet.cell(row = j, column = AMOUNT_REVENUE_COLUMN) 
                total_revenue_income += amount_cell.value


        # total service
        sheet.cell(row = i, column = 2).value = "Total Service"
        sheet.cell(row = i, column = 2).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_labor_cost
        sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = 0.0
        sheet.cell(row = i, column = DIFF_COLUMN).value = 0.0
        i += 1

        # total income
        sheet.cell(row = i, column = 2).value = "Total Income"
        sheet.cell(row = i, column = 2).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = 0.0
        sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = total_revenue_income
        sheet.cell(row = i, column = DIFF_COLUMN).value = 0.0
        i += 1

        # total
        sheet.cell(row = i, column = 1).value = "Total"
        sheet.cell(row = i, column = 1).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_cost
        sheet.cell(row = i, column = ACT_REVENUE_COLUMN).value = total_revenue_income
        sheet.cell(row = i, column = DIFF_COLUMN).value = total_revenue_income - total_cost
        # total font bold
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_REVENUE_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = DIFF_COLUMN).font = Font(bold=True)

        i += 1
        # whitespace
        i += 1
    
        # summary box
        '''
        Calculation details for the summary box:
        Total Labor: Add all labor costs except temp labor
        Labor OH: Multiply 30% to the total labor calculated above
        Other OH: Multiply 0.005 to all costs
        Total Cost w/OH: Total Costs + Labor OH + Other OH
        '''
        labor_oh = total_labor_cost * 0.3
        other_oh = total_cost * 0.005
        total_cost_w_oh = total_cost + labor_oh + other_oh

        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total Labor"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value =  total_labor_cost
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Labor OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = labor_oh
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Other OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = other_oh
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total Cost w/ OH"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_cost_w_oh
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Billed To Date"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_revenue_income
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1

        # column letter row number : column letter row number  for top left, bottom right
        cell_range = "D" + str(i-5) + ":E" + str(i-1)
        set_border(sheet, cell_range)

        # whitespace
        i += 1

        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Billed To Date"
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        i += 1

       # Grab all billed
       # Could avoid doing this iteration twice
        for j in range(6, revenue_sheet.max_row + 1):    
            j_name = revenue_sheet.cell(row = j, column = NAME_REVENUE_COLUMN).value
            if j_name and job_number in j_name:

                memo_cell = revenue_sheet.cell(row = j, column = MEMO_REVENUE_COLUMN) 
                item_cell = revenue_sheet.cell(row = j, column = ITEM_REVENUE_COLUMN) 
                amount_cell = revenue_sheet.cell(row = j, column = AMOUNT_REVENUE_COLUMN) 

                sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = memo_cell.value
                sheet.cell(row = i, column = ACT_COST_COLUMN).value = amount_cell.value
                i += 1

        # write total income for the last time
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).value = "Total" 
        sheet.cell(row = i, column = SUBITEM_NAME_COLUMN).font = Font(bold=True)
        sheet.cell(row = i, column = ACT_COST_COLUMN).value = total_revenue_income
        sheet.cell(row = i, column = ACT_COST_COLUMN).font = Font(bold=True)
        i += 1

        # write date range
        sheet.cell(row = 3, column = 1).value = "Transactions from: " + min_date.strftime("%m/%d/%y") + " to " + max_date.strftime("%m/%d/%y")

        # clear out extra rows
        sheet.delete_rows(i, sheet.max_row - i)

        # trim printable area to data?
        sheet._print_area = "A1:I"+str(i)

    # -------------------------------------------------------------------------------- #

    # create and fill all job sheet data
    for sheet in cost_detail_wb:
        if sheet.title == 'Total':
            continue
        # copy initial format into empty sheet
        copySheet(jc_wb.active, sheet)
        createJobCostSheet(sheet)

    cost_detail_wb.save(processed_file_path)


def main(argv):
    if len(argv) == 0 or len(argv) > 2:
        print("Error - usage: supply cost detail workbook , revenue workbook")
        return

    cost_detail_wb_path = os.path.abspath(argv[0])
    revenue_wb_path     = os.path.abspath(argv[1])

    if os.path.isfile(cost_detail_wb_path) and os.path.isfile(revenue_wb_path):
        createJobWorkbook(cost_detail_wb_path, revenue_wb_path)
    elif not os.path.isfile(cost_detail_wb_path):
        print("Error: cost detail workbook does not exist?: ", cost_detail_wb_path)
    elif not os.path.isfile(revenue_wb_path):
        print("Error: revenue workbook does not exist?: ", revenue_wb_path)
    else:
        print("Error: wrong input")

if __name__ == "__main__":
   main(sys.argv[1:])
