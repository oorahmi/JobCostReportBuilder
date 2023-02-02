'''
Program to take in a eva report and output a wip report

Usage:  

    python createWIPReport.py path_to_eva_workbook_file quarter_to_process(1-4) year

The wip report will be saved in /processed

Requires openpyxl

    pip install openpyxl

Author: Brian Wright
10-19-2022

'''
import sys
import openpyxl
import os 
import shutil
from copy import copy
import datetime
from collections import OrderedDict
       

def createWIPReport(eva_wb_path, current_quarter, year):
    eva_total_wb = openpyxl.load_workbook(eva_wb_path) 
    if not eva_total_wb:
        print("Error: failed to open workbook: ", eva_wb_path)
        return
    
    # TODO: pull current quarter from eva report?
    if current_quarter < 1 or current_quarter > 4:
        print("Error: Input valid quarter 1-4")
        return
    
    # copy empty wip
    wip_blank_path = os.getcwd() + "/data/wip_blank.xlsx"
    processed_file_path = os.getcwd() +'/processed/wip_processed.xlsx'
    shutil.copyfile(wip_blank_path, processed_file_path)

    wip_report_wb = openpyxl.load_workbook(wip_blank_path) 
    if not wip_report_wb:
        print("Error: failed to open data workbook: ", wip_blank_path)
        return

    actual_cost_detail_sheet    = eva_total_wb.worksheets[0]
    # revenue_sheet               = eva_total_wb.worksheets[1]
    estimate_cost_detail_sheet  = eva_total_wb.worksheets[2]

    wip_report_sheet = wip_report_wb.worksheets[0]

    min_date = datetime.date.max
    max_date = datetime.date.min

    if current_quarter == 1:
        min_date = datetime.datetime(year, 1, 1, 0, 0, 0)
        max_date = datetime.datetime(year, 3, 31, 0, 0, 0)
    elif current_quarter == 2:
        min_date = datetime.datetime(year, 4, 1, 0, 0, 0)
        max_date = datetime.datetime(year, 6, 30, 0, 0, 0)
    elif current_quarter == 3:
        min_date = datetime.datetime(year, 7, 1, 0, 0, 0)
        max_date = datetime.datetime(year, 9, 30, 0, 0, 0)
    elif current_quarter == 4:
        min_date = datetime.datetime(year, 10, 1, 0, 0, 0)
        max_date = datetime.datetime(year, 12, 31, 0, 0, 0)
    
    min_date.today()

    # define all columns
    WIP_JOB_NUMBER_COlUMN = 1
    WIP_JOB_NAME_COlUMN = 2
    WIP_CONTRACT_PRICE_COlUMN = 4
    WIP_APPROVED_CO_COlUMN = 5
    WIP_TOTAL_PRICE_COlUMN = 6
    WIP_ESTIMATED_COST_COlUMN = 7
    WIP_ESTIMATED_PROFIT_COlUMN = 8
    WIP_ESTIMATED_PROFIT_PERC_COlUMN = 9
    WIP_ACTUAL_COST_TO_DATE_COlUMN = 10
    WIP_PERC_COMPLETION_COlUMN = 11
    WIP_REVENUES_EARNED_TO_DATE_COlUMN = 12
    WIP_BILLINGS_TO_DATE_COlUMN = 13
    WIP_RETAINAGE_COlUMN = 14
    WIP_COST_IN_EXEC_BILLINGS_COlUMN = 15
    WIP_BILLINGS_IN_EXCESS_COlUMN = 16
    WIP_BACKLOG_COlUMN = 17
    WIP_Q_REVENUES_COlUMN = 18 # NOTE: omitting calculation
    WIP_Q_COSTS_COlUMN = 19
    WIP_Q_LAB_OVERHEAD_COlUMN = 20
    WIP_Q_OTH_OVERHEAD_COlUMN = 21
    WIP_Q_PROFIT_LOSS_COlUMN = 22

    # do quarter dates for columns
    wip_report_sheet.cell(row = 3, column = 1).value = "As of " + str(current_quarter) + "Q" + str(year)
    wip_report_sheet.cell(row = 5, column = WIP_Q_REVENUES_COlUMN).value = str(current_quarter) + "Q" + str(year)
    wip_report_sheet.cell(row = 5, column = WIP_Q_COSTS_COlUMN).value = str(current_quarter) + "Q" + str(year)
    wip_report_sheet.cell(row = 5, column = WIP_Q_LAB_OVERHEAD_COlUMN).value = str(current_quarter) + "Q" + str(year)
    wip_report_sheet.cell(row = 5, column = WIP_Q_OTH_OVERHEAD_COlUMN ).value = str(current_quarter) + "Q" + str(year)
    wip_report_sheet.cell(row = 5, column = WIP_Q_PROFIT_LOSS_COlUMN ).value = str(current_quarter) + "Q" + str(year)
    wip_report_sheet.title = str(current_quarter) + "Q" + str(year)


    if len(eva_total_wb.worksheets)  <= 3:
        print("Error: eva workbook does not have job sheets?")
        return 

    fl_job_numbers = []
    m_fl_job_numbers = []
    other_job_numbers = []
    # find order of jobs
    # get all food lion jobs first, sort ascending, then rest sort ascending
    for i in range(3, len(eva_total_wb.worksheets)):
        job_sheet = eva_total_wb.worksheets[i]
        job_desc = job_sheet.cell(row = 2, column = 1).value
        job_number = job_sheet.title
        if "Food Lion" in job_desc:
            if "Maintenance" in job_desc:
                m_fl_job_numbers.append(job_number)
            else:
                fl_job_numbers.append(job_number)
        else:
            other_job_numbers.append(job_number)

    fl_job_numbers.sort()
    m_fl_job_numbers.sort()
    other_job_numbers.sort()
    
    # build sorted list from ordered job numbers
    ordered_job_numbers = fl_job_numbers + m_fl_job_numbers + other_job_numbers

    # NOTE: dont need to store these seperately but nicer to work with?
    ordered_sheets = []
    for j_number in ordered_job_numbers:
        # get correct sheet
        for i in range(3, len(eva_total_wb.worksheets)): 
            job_sheet = eva_total_wb.worksheets[i]
            if j_number == job_sheet.title:
                ordered_sheets.append(job_sheet)
    
    # write job data
    i = 7 # after header
    for job_sheet in ordered_sheets:
        
        k = i # tmp for row
        i = str(i)
        # write formulaic columns
        wip_report_sheet.cell(row = k, column = WIP_TOTAL_PRICE_COlUMN).value = "=D" + i + "+E" + i
        wip_report_sheet.cell(row = k, column = WIP_ESTIMATED_PROFIT_COlUMN).value = "=+F" + i + "-G" + i
        wip_report_sheet.cell(row = k, column = WIP_ESTIMATED_PROFIT_PERC_COlUMN).value = "=IFERROR((H" + i + "/G" + i + "), 0)"
        wip_report_sheet.cell(row = k, column = WIP_PERC_COMPLETION_COlUMN).value = "=IFERROR((J" + i + "/G" + i + "), 0)"
        wip_report_sheet.cell(row = k, column = WIP_REVENUES_EARNED_TO_DATE_COlUMN).value = "=F" + i + "*K" + i
        wip_report_sheet.cell(row = k, column = WIP_COST_IN_EXEC_BILLINGS_COlUMN).value = '=IF(L'+i+'>M'+i+',L'+i+'-M'+i+',IF((L'+i+'-M'+i+')<-1,"-",0))'
        wip_report_sheet.cell(row = k, column = WIP_BILLINGS_IN_EXCESS_COlUMN).value = '=IF(L'+i+'>M'+i+',"-",L'+i+'-M'+i+')'
        wip_report_sheet.cell(row = k, column = WIP_BACKLOG_COlUMN).value = "=+F"+i+"-L"+i
        wip_report_sheet.cell(row = k, column = WIP_Q_PROFIT_LOSS_COlUMN).value = "=R"+i+"-S"+i
        i = int(i)

        job_number = job_sheet.title

        job_name = job_sheet.cell(row = 2, column = 1).value
        # trimming off "Job Estimates vs. Actuals Detail for "
        job_name = job_name[37:]
        name_index = job_name.find(job_number) + 7 # job number length
        job_text = job_name[name_index:].strip()
        if "Food Lion" in job_name:
            if job_number in m_fl_job_numbers:
                wip_report_sheet.cell(row = i, column = WIP_JOB_NAME_COlUMN).value = "Maint. FL " + str(job_number[3:]) + " " + job_text
                wip_report_sheet.cell(row = i, column = WIP_JOB_NUMBER_COlUMN).value = "M" + job_number
            else:
                # only write 4 digits of the job number and the rest of the text.
                wip_report_sheet.cell(row = i, column = WIP_JOB_NAME_COlUMN).value = "FL " + str(job_number[3:]) + " " + job_text
                wip_report_sheet.cell(row = i, column = WIP_JOB_NUMBER_COlUMN).value = job_number
        else:
            # only write the rest of the text after job number 
            wip_report_sheet.cell(row = i, column = WIP_JOB_NAME_COlUMN).value = job_text
            wip_report_sheet.cell(row = i, column = WIP_JOB_NUMBER_COlUMN).value = job_number


        JOB_ESTIMATE_COLUMN = 5
        estimate_total = 0
        change_order = 0
        orig_contract = 0
        other_job_income = 0
        billed_to_date = 0
        total_cost_w_oh = 0
        retainage = 0

        labor_oh = 0
        other_oh = 0


        # get several numbers from job sheet
        for j in range(7, job_sheet.max_row + 1): # could optimize by not doing all rows

            if job_sheet.cell(row = j, column = 2).value == "Total":
                estimate_total = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 4).value == "Orig Contract":
                orig_contract= job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 4).value == "Change Order":
                change_order = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 4).value == "Other Job Income":
                other_job_income = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 3).value == "Total Billed to Date":
                billed_to_date = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 4).value == "Total Cost w/ OH":
                total_cost_w_oh = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 3).value == "Retainage Held by Customer":
                retainage = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            # maybe unnecessary
            elif job_sheet.cell(row = j, column = 4).value == "Labor OH":
                labor_oh = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value
            elif job_sheet.cell(row = j, column = 4).value == "Other OH":
                other_oh = job_sheet.cell(row = j, column = JOB_ESTIMATE_COLUMN).value

        
        ESTIMATE_NAME_COLUMN   = 10
        ESTIMATE_DESC_COLUMN   = 14
        ESTIMATE_AMOUNT_COLUMN = 16

        #if larger than 
        if estimate_total > 0:
            wip_report_sheet.cell(row = i, column = WIP_CONTRACT_PRICE_COlUMN).value = estimate_total 

            # need to factor in extra original cost if applicable
            approved_co = 0
            if orig_contract > estimate_total:
                approved_co = (orig_contract - estimate_total) + change_order + other_job_income
            else:
                approved_co = change_order + other_job_income

            wip_report_sheet.cell(row = i, column = WIP_APPROVED_CO_COlUMN).value = approved_co

            # find orig_contract income from estimate sheet, only affets
            orig_contract_from_estimate = 0

            for t in range(5, estimate_cost_detail_sheet.max_row + 1):    # could optimize by not doing all rows
                e_name = estimate_cost_detail_sheet.cell(row = t, column = ESTIMATE_NAME_COLUMN).value
                e_desc = estimate_cost_detail_sheet.cell(row = t, column = ESTIMATE_DESC_COLUMN).value
                if not e_name or not e_desc:
                    continue
                if job_number in e_name and "Original Contract Income" == e_desc:
                    e_amount = estimate_cost_detail_sheet.cell(row = t, column = ESTIMATE_AMOUNT_COLUMN).value
                    orig_contract_from_estimate += e_amount
            

            wip_report_sheet.cell(row = i, column = WIP_ESTIMATED_COST_COlUMN).value = (estimate_total - orig_contract_from_estimate) + (approved_co * .95)


        else: # no estimate
            wip_report_sheet.cell(row = i, column = WIP_CONTRACT_PRICE_COlUMN).value = orig_contract 
            wip_report_sheet.cell(row = i, column = WIP_APPROVED_CO_COlUMN).value = change_order + other_job_income
            #                                                                           not sure if change_order makes sense here, but just need previous total price
            wip_report_sheet.cell(row = i, column = WIP_ESTIMATED_COST_COlUMN).value = (orig_contract + change_order + other_job_income ) * .95


        # fill in simple columns
        wip_report_sheet.cell(row = i, column = WIP_ACTUAL_COST_TO_DATE_COlUMN).value = total_cost_w_oh 
        wip_report_sheet.cell(row = i, column = WIP_BILLINGS_TO_DATE_COlUMN).value = billed_to_date 
        wip_report_sheet.cell(row = i, column = WIP_RETAINAGE_COlUMN).value = retainage 

        # quarterly handling
        ACTUAL_NAME_COLUMN = 11 
        ACTUAL_ITEM_COLUMN = 15      
        ACTUAL_AMOUNT_COLUMN = 19
        ACTUAL_DATE_COLUMN = 9

        total_labor_cost_no_temp = 0
        total_costs = 0

        for k in range(5, actual_cost_detail_sheet.max_row + 1):    # could optimize by not doing all rows
            j_name = actual_cost_detail_sheet.cell(row = k, column = ACTUAL_NAME_COLUMN).value
            j_date = actual_cost_detail_sheet.cell(row = k, column = ACTUAL_DATE_COLUMN).value
            if not j_name or not j_date:
                continue

            # looking at the relevant job and within the same quarter?
            if job_number in j_name and j_date >= min_date and j_date <= max_date:
                job_item_desc = actual_cost_detail_sheet.cell(row = k, column = ACTUAL_ITEM_COLUMN).value
                job_item_desc = job_item_desc.lower()
                j_amount = actual_cost_detail_sheet.cell(row = k, column = ACTUAL_AMOUNT_COLUMN).value
                total_costs += j_amount
                if "labor" in job_item_desc and "temp" not in job_item_desc:
                    total_labor_cost_no_temp += j_amount

        '''
        Calculation details for the summary box:
        Total Labor: Add all labor costs except temp labor
        Labor OH: Multiply 30% to the total labor calculated above
        Other OH: Multiply 0.005 to all costs
        Total Cost w/OH: Total Costs + Labor OH + Other OH
        '''
        q_labor_oh = total_labor_cost_no_temp * 0.3
        q_other_oh = total_costs * 0.005
        q_total_cost_w_oh = total_costs + q_labor_oh + q_other_oh

        wip_report_sheet.cell(row = i, column = WIP_Q_COSTS_COlUMN).value = q_total_cost_w_oh 
        wip_report_sheet.cell(row = i, column = WIP_Q_LAB_OVERHEAD_COlUMN).value = q_labor_oh
        wip_report_sheet.cell(row = i, column = WIP_Q_OTH_OVERHEAD_COlUMN).value = q_other_oh

        i += 1 # next job
    
    # calculate totals of each column
    # writing to bottom of current bounding box
    # TODO: dynamic bounding box and do this after all jobs
    wip_report_sheet.cell(row = 61, column = WIP_CONTRACT_PRICE_COlUMN).value ='=SUM(D7:D60)' 
    wip_report_sheet.cell(row = 61, column = WIP_APPROVED_CO_COlUMN).value ='=SUM(E7:E60)'  
    wip_report_sheet.cell(row = 61, column = WIP_TOTAL_PRICE_COlUMN).value = '=SUM(F7:F60)' 
    wip_report_sheet.cell(row = 61, column = WIP_ESTIMATED_COST_COlUMN).value ='=SUM(G7:G60)'
    wip_report_sheet.cell(row = 61, column = WIP_ESTIMATED_PROFIT_COlUMN).value = '=SUM(H7:H60)' 
    wip_report_sheet.cell(row = 61, column = WIP_ESTIMATED_PROFIT_PERC_COlUMN).value = '=AVERAGE(I7:I60)' 
    wip_report_sheet.cell(row = 61, column = WIP_ACTUAL_COST_TO_DATE_COlUMN).value = '=SUM(J7:J60)' 
    wip_report_sheet.cell(row = 61, column = WIP_PERC_COMPLETION_COlUMN).value = '=AVERAGE(K7:K60)' 
    wip_report_sheet.cell(row = 61, column = WIP_REVENUES_EARNED_TO_DATE_COlUMN).value = '=SUM(L7:L60)' 
    wip_report_sheet.cell(row = 61, column = WIP_BILLINGS_TO_DATE_COlUMN).value = '=SUM(M7:M60)' 
    wip_report_sheet.cell(row = 61, column = WIP_RETAINAGE_COlUMN).value = '=SUM(N7:N60)' 
    wip_report_sheet.cell(row = 61, column = WIP_COST_IN_EXEC_BILLINGS_COlUMN).value = '=SUM(O7:O60)' 
    wip_report_sheet.cell(row = 61, column = WIP_BILLINGS_IN_EXCESS_COlUMN).value = '=SUM(P7:P60)' 
    wip_report_sheet.cell(row = 61, column = WIP_BACKLOG_COlUMN).value = '=SUM(Q7:Q60)' 
    wip_report_sheet.cell(row = 61, column = WIP_Q_REVENUES_COlUMN).value = '=SUM(R7:R60)' 
    wip_report_sheet.cell(row = 61, column = WIP_Q_COSTS_COlUMN).value = '=SUM(S7:S60)' 
    wip_report_sheet.cell(row = 61, column = WIP_Q_LAB_OVERHEAD_COlUMN).value = '=SUM(T7:T60)' 
    wip_report_sheet.cell(row = 61, column = WIP_Q_OTH_OVERHEAD_COlUMN).value = '=SUM(U7:U60)' 
    wip_report_sheet.cell(row = 61, column = WIP_Q_PROFIT_LOSS_COlUMN).value = '=SUM(V7:V60)' 

    wip_report_wb.save(processed_file_path)

    # -------------------------------------------------------------------------------- #


def main(argv):
    if len(argv) == 0 or len(argv) > 3:
        print("Error - usage: supply the processed EVA Job Workbook")
        return

    eva_wb_path = os.path.abspath(argv[0])
    quarter = int(argv[1])
    year = int(argv[2])

    if os.path.isfile(eva_wb_path):
        createWIPReport(eva_wb_path, quarter, year) 
    elif not os.path.isfile():
        print("Error: eva workbook path does not exist?: ", eva_wb_path)
    else:
        print("Error: wrong input")

if __name__ == "__main__":
   main(sys.argv[1:])
